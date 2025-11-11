# Import libraries
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

#Enter report type you want to retrieve (B1 or B3)
report_type = 'B3'

# Create dataframes from csv files
collection_df = pd.read_csv('deselection_collection.csv')
usage_df = pd.read_csv('deselection_usage_'+ report_type +'.csv', dtype={'Normalized ISBN': str})
overlap_df = pd.read_excel('Compare Collections - Titles of portfolios with complete coverage overlap.xlsx',
                           sheet_name='Titles with complete overlap')

# Remove unwanted characters from portfolio column from overlap analysis
overlap_df['Portfolio ID - Source Set'] = overlap_df['Portfolio ID - Source Set'].str.strip('{"}')

# Define the expected years
expected_years = [2020, 2021, 2022, 2023, 2024, 2025]

# Create pivot table from usuage dataframe
pivot_table = usage_df.pivot_table(
    values='TR_'+ report_type + ' - Total Item Requests',
    index=['Interface', 'Normalized Title', 'Normalized ISBN', 'Year Of Publication'],
    columns='Usage Date Year',
    aggfunc='sum',
    fill_value=0,
    margins=True,  # Adds the total column to the right
    margins_name='Total'  # Name total column
).reset_index()

# Remove the total row from the end of the pivot table
pivot_table = pivot_table[pivot_table['Interface'] != 'Total']

# Ensure all expected years are present in the pivot_table, adding missing ones filled with 0
for year in expected_years:
    if year not in pivot_table.columns:
        pivot_table[year] = 0

# Reorder columns to ensure the years are in the correct order
pivot_table = pivot_table[['Interface', 'Normalized Title', 'Normalized ISBN', 'Year Of Publication'] 
                          + expected_years + ['Total']]

# Create dictionary from usuage pivot table with usuage for each year
isbn_to_years = dict(zip(pivot_table['Normalized ISBN'], zip(pivot_table[2020], pivot_table[2021], pivot_table[2022], 
                                                             pivot_table[2023], pivot_table[2024], pivot_table[2025], 
                                                             pivot_table['Total'])))

# Function to match usage ISBNs with Collection ISBNs and add usage data to collection dataframe
def match_and_copy_usage(isbn_list):
    # Ensure that the isbn_list is treated as a string
    isbn_list = str(isbn_list)
    
    # Iterate through each ISBN in the dictionary and check if it's contained in the isbn_list string
    for isbn in isbn_to_years.keys():
        # Ensure that ISBN from the dictionary is also treated as a string for comparison
        if str(isbn) in isbn_list:
            # Return the tuple of year values as integers (converting np.nan to 0 if needed)
            return tuple(int(value) if not np.isnan(value) else 0 for value in isbn_to_years[isbn])
    
    # Return 0 tuple if no match found.
    return (0, 0, 0, 0, 0, 0, 0)

# Apply function to update collection_df columns
matched_data = collection_df['ISBN (Normalized)'].apply(match_and_copy_usage)

# Only assign values where there was a match, ensure integers for the year columns
collection_df[['2020', '2021', '2022', '2023', '2024', '2025',
               'Total']] = pd.DataFrame(matched_data.tolist(), index=matched_data.index)

# Create dictionary from overlap analysis with portfolio and with the collections with overlap.
portfolios_ID_overlap = dict(zip(overlap_df['Portfolio ID - Source Set'], zip(overlap_df['Collection - Target Set'])))

# Function to match overlap portoflio ids with Collection portoflio ids and add overlap info to collection dataframe.
def match_and_copy_overlap(portfolio_list):
    # Ensure that the portfolio_list is treated as a string
    portfolio_list = str(portfolio_list)
    
    # Iterate through each portfolio in the dictionary and check if it's contained in the portfolio_list string
    for portfolio in portfolios_ID_overlap.keys():
        # Ensure that portfolio from the dictionary is also treated as a string for comparison
        if str(portfolio) in portfolio_list:
            # Return the tuple of values
            return portfolios_ID_overlap[portfolio]
    
    # Return 'no overlap' if no match is found
    return ('No overlap',)

# Apply function to update collection_df columns
matched_data = collection_df['Portfolio Id'].apply(match_and_copy_overlap)

# Change ID values to strings so they will display correctly in excel export.
collection_df[['Overlap collection(s)']] = pd.DataFrame(matched_data.tolist(), index=collection_df.index)
collection_df[['Electronic Collection Id','Portfolio Id']] = collection_df[['Electronic Collection Id','Portfolio Id']].astype(str)
overlap_df[['MMS ID']] = overlap_df[['MMS ID']].astype(str)
usage_df[['Normalized ISBN']] = usage_df[['Normalized ISBN']].astype(str)
pivot_table[['Normalized ISBN']] = pivot_table[['Normalized ISBN']].astype(str)

# Create a Pandas Excel writer using openpyxl, add dataframes to sheets.
with pd.ExcelWriter('deselection_output.xlsx', engine='openpyxl') as writer:
    collection_df.to_excel(writer, index=False, sheet_name='Compiled data')
    pivot_table.to_excel(writer, index=False, sheet_name='Usage pivot table')
    usage_df.to_excel(writer, index=False, sheet_name='Usage data')
    overlap_df.to_excel(writer, index=False, sheet_name='Overlap analysis')
    

    # Access the workbook and the sheet
    workbook = writer.book
    worksheet = writer.sheets['Compiled data']
    
    # Get the number of rows in the DataFrame (excluding the header)
    num_rows = len(collection_df)

    # Define the red fill color
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Format cells in columns J to O where the value in column O is 0
    for row in worksheet.iter_rows(min_row=2, max_row=num_rows+1, min_col=11, max_col=17):  # Columns K to P
        if row[-1].value == 0:  # Check if the last cell in the row (column O) is P
            for cell in row:
                cell.fill = red_fill  # Fill cells with red color
                
                
